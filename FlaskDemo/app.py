from flask import Flask, render_template, request, make_response, redirect, url_for
import pandas as pd
import shutil
import tempfile
import zipfile
from markupsafe import Markup
import os
import openpyxl

app = Flask(__name__)

base_dir = os.path.abspath(os.path.dirname(__file__))
file_path = os.path.join(base_dir, 'sampledata.xlsx')
print(f"critical: writing to -> {os.path.abspath(file_path)}")
#dfs = pd.read_excel(file_path, sheet_name=None)

studentSheet = pd.read_excel(file_path, sheet_name="student answers")
studentSheet['Answer'] = (
    studentSheet['Answer'].fillna('').str.replace('\n', '<br>', regex=False)
)


# rubricSheet = dfs["rubric"]
# questionSheet = dfs["question versions"]



#-------- Trying to make tables static this calling them outside of the routing and main flask as they dont need to change

# Reading only the rubric and question frames to avoid these static tables being replaced when writting into pages
rubric_df = pd.read_excel(file_path, sheet_name="rubric").fillna('').replace(r'\n', '<br>', regex=True)
STATIC_RUBRIC_HTML = Markup(rubric_df.to_html(index=False, border=0, escape=False))
del rubric_df

question_df = pd.read_excel(file_path, sheet_name="question versions").fillna('').replace(r'\n', '<br>', regex=True)
question_df = question_df.rename(columns={"Unnamed: 0": " "})
STATIC_QUESTION_HTML = Markup(question_df.to_html(index=False, border=0, escape=False))
del question_df

#--------

RUBRIC_COLUMNS = []  

@app.route("/upload", methods=["GET", "POST"])
def upload_file():
    global file_path, studentSheet, STATIC_RUBRIC_HTML, STATIC_QUESTION_HTML, RUBRIC_COLUMNS
    
    if request.method == "POST":
        file = request.files.get("file")
        if file and file.filename.endswith(".xlsx"):
            
            base_dir = os.path.abspath(os.path.dirname(__file__)) 
            file_path = os.path.join(base_dir, 'sampledata.xlsx') 
            file.save(file_path)
            
            
            studentSheet = pd.read_excel(file_path, sheet_name="student answers")
            studentSheet['Answer'] = studentSheet['Answer'].fillna('').str.replace('\n', '<br>', regex=False)
            
            
            RUBRIC_COLUMNS = [col for col in studentSheet.columns if str(col).startswith("Teacher Score")]
            
           
            for col in RUBRIC_COLUMNS:
                studentSheet[col] = studentSheet[col].astype(float)
                
            rubric_df = pd.read_excel(file_path, sheet_name="rubric").fillna('').replace(r'\n', '<br>', regex=True)
            STATIC_RUBRIC_HTML = Markup(rubric_df.to_html(index=False, border=0, escape=False))
            del rubric_df 

            question_df = pd.read_excel(file_path, sheet_name="question versions").fillna('').replace(r'\n', '<br>', regex=True)
            question_df = question_df.rename(columns={"Unnamed: 0": " "})
            STATIC_QUESTION_HTML = Markup(question_df.to_html(index=False, border=0, escape=False))
            del question_df
            
            return redirect(url_for("home"))
            
    return render_template (
        "upload.html"
    )


@app.route("/", methods=["GET", "POST"])
def home():
    global studentSheet, RUBRIC_COLUMNS
    message = "" 
    
    if studentSheet.empty:
        return "Please upload an Excel file first at <a href='/upload'>/upload</a>"
    
    #If RUBRIC_COLUMNS became empty due to a server refresh, re-detect them
    if not RUBRIC_COLUMNS:
        RUBRIC_COLUMNS = [col for col in studentSheet.columns if str(col).startswith("Teacher Score")]

    if request.method == "POST": 
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['student answers']
        
        col_indices = {col: studentSheet.columns.get_loc(col) + 1 for col in RUBRIC_COLUMNS}
        
        for index in studentSheet.index: 
            excel_row = int(index) + 2
            
            for col_name in RUBRIC_COLUMNS:
                val = request.form.get(f"{col_name}_{index}", "").strip() 
                if val:
                    val_float = float(val)
                    studentSheet.loc[index, col_name] = val_float 
                    ws.cell(row=excel_row, column=col_indices[col_name], value=val_float)
        
        wb.save(file_path)
        wb.close()
        message = "All scores saved successfully." 


    students = []

    
    for idx, row in studentSheet.iterrows():
        record = row.to_dict()
        record["row_index"] = idx
        students.append(record)

    return render_template(
        "table.html",
        students=students,
        message=message,
        rubric_columns=RUBRIC_COLUMNS
    )

@app.route("/student/<int:index>", methods=["GET", "POST"])
def student(index):
    global STATIC_RUBRIC_HTML, STATIC_QUESTION_HTML, studentSheet, RUBRIC_COLUMNS
    message = "" 
    
    if studentSheet.empty:
        return "Please upload Excel file first at <a href='/upload'>/upload</a>"
    
    if request.method == "POST": 
        wb = openpyxl.load_workbook(file_path, data_only=False)
        ws = wb['student answers']
        
        excel_row = int(index) + 2
        
        # Loop dynamically
        for col_name in RUBRIC_COLUMNS:
            if col_name in request.form:
                val_float = float(request.form[col_name])
                studentSheet.loc[index, col_name] = val_float
                
                col_idx = studentSheet.columns.get_loc(col_name) + 1
                ws.cell(row=excel_row, column=col_idx, value=val_float)
        
        wb.save(file_path)
        wb.close()
        message = "Scores saved successfully." 
        
    student = studentSheet.iloc[index] 


    return render_template(
        "student.html",
        student=student,
        #student_table=Markup(student_HTML),
        rubric_table=STATIC_RUBRIC_HTML,
        question_table=STATIC_QUESTION_HTML,
        index=index,
        total=len(studentSheet),
        message=message,
        rubric_columns=RUBRIC_COLUMNS
    )

if __name__ == "__main__":
    app.run(debug=True)